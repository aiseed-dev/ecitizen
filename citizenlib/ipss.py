"""国立社会保障・人口問題研究所(IPSS)「日本の地域別将来推計人口(令和5(2023)年推計)」
の都道府県別 Excel ファイルをパースする。

出典: https://www.ipss.go.jp/pp-shicyoson/j/shicyoson23/3kekka/Municipalities.asp
      (47都道府県別、各 Municipalities/{01..47}.xlsx)
ダウンロードは 1 回限り (`tools/fetch_ipss.py`)。data/raw/ipss/ にキャッシュし、
再生成可能なので git 管理外 (.gitignore)。

シート構成: 1 シート = 1 自治体 (シート名は "{4桁コード}_{名称}"。
都道府県計は先頭シートで "{2桁コード}00_{都道府県名}")。
列構成: 0-7=男女計(ラベル+2020,2025,2030,2035,2040,2045,2050年)、8=区切り、
        9-16=男、17=区切り、18-25=女。
2020年列は「国勢調査による実績値」(推計ではない)。
年齢区分は 90～94歳・95歳～ の2区分 (旧 Ages3 の「90歳以上」より細かい)。
"""
import re
from pathlib import Path

import openpyxl

from . import masters

ROOT = Path(__file__).resolve().parent.parent
IPSS_DIR = ROOT / "data" / "raw" / "ipss"

# シート内の年齢階級行 (この順で総数の次に並ぶ。総数・（再掲）行は別途扱う)
AGE_ROWS_IPSS = [
    "0～4歳", "5～9歳", "10～14歳", "15～19歳", "20～24歳", "25～29歳", "30～34歳",
    "35～39歳", "40～44歳", "45～49歳", "50～54歳", "55～59歳", "60～64歳",
    "65～69歳", "70～74歳", "75～79歳", "80～84歳", "85～89歳", "90～94歳", "95歳～",
]
PROJECTION_YEARS = [2020, 2025, 2030, 2035, 2040, 2045, 2050]

_BLOCK_COLS = {"total": 0, "male": 9, "female": 18}  # 各ブロックのラベル列オフセット


def _parse_sheet(ws) -> dict:
    """1シート分をパースし、{block: [{"series":..,"population":[7]}, ...21行]} を返す。

    21行は旧 Ages3 互換 (総数, 0～4歳, ..., 85～89歳, 90歳以上, 年齢不詳)。
    90～94歳+95歳～ を合算して「90歳以上」に、年齢不詳は 0 とする
    (IPSS 推計に年齢不詳区分は無い)。
    """
    rows_by_label = {}
    for row in ws.iter_rows(values_only=True):
        for block, col in _BLOCK_COLS.items():
            label = row[col]
            if label is None:
                continue
            label = str(label).strip().lstrip("(（").rstrip(")）")  # "（再掲）" 等の全角括弧除去用の軽い正規化
            rows_by_label.setdefault(block, {})[label] = list(row[col + 1:col + 8])

    out = {}
    for block in ("total", "male", "female"):
        by_label = rows_by_label[block]
        series = [{"series": "総数", "population": by_label["総数"]}]
        for age in AGE_ROWS_IPSS[:18]:  # 0～4歳 .. 85～89歳
            series.append({"series": age, "population": by_label[age]})
        merged90 = [a + b for a, b in zip(by_label["90～94歳"], by_label["95歳～"])]
        series.append({"series": "90歳以上", "population": merged90})
        series.append({"series": "年齢不詳", "population": [0] * len(PROJECTION_YEARS)})
        out[block] = series
    return out


class IpssData:
    """47都道府県分の xlsx を必要に応じて読み込み、キャッシュする。"""

    def __init__(self, ipss_dir: Path = IPSS_DIR):
        self.ipss_dir = ipss_dir
        self._workbooks = {}

    def _wb(self, pref: str):
        if pref not in self._workbooks:
            path = self.ipss_dir / f"{pref}.xlsx"
            if not path.exists():
                raise FileNotFoundError(
                    f"{path} が見つかりません。tools/fetch_ipss.py で先にダウンロードしてください。")
            self._workbooks[pref] = openpyxl.load_workbook(path, data_only=True, read_only=True)
        return self._workbooks[pref]

    def sheet_codes(self, pref: str) -> set:
        """この都道府県ファイルに含まれる5桁市区町村コード一覧 (先頭の都道府県計シートを除く)。"""
        wb = self._wb(pref)
        return {re.match(r"^\d+", s).group().zfill(5) for s in wb.sheetnames[1:]}

    def city(self, code: str) -> dict | None:
        """市区町村コード(5桁)のデータ。シートが無ければ None (例: 福島県浜通り13町村)。

        CITY_DIC (2016-10-10時点) 以降の市制施行等によるコード変更は
        masters.IPSS_CODE_TRANS で吸収する。
        """
        code = masters.IPSS_CODE_TRANS.get(code, code)
        pref = code[:2]
        wb = self._wb(pref)
        code4 = str(int(code))  # "01100" -> "1100"
        for name in wb.sheetnames[1:]:
            if re.match(r"^\d+", name).group() == code4:
                return _parse_sheet(wb[name])
        return None

    def prefecture(self, pref: str) -> dict:
        """都道府県計 (各ファイルの先頭シート)。"""
        wb = self._wb(pref)
        return _parse_sheet(wb[wb.sheetnames[0]])

    def close(self) -> None:
        for wb in self._workbooks.values():
            wb.close()
        self._workbooks.clear()
