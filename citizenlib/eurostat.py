"""Eurostat (census: demo_pjangroup) / ONS (UK projection: 2024-based Principal
projection) から取得した Country(海外)ページ用データをパースする。

キャッシュ元:
  data/raw/eurostat/census.json     (tools/fetch_eurostat.py)
  data/raw/eurostat/projection.json (同上、UK は対象外)
  data/raw/ons/uk_ppp_machine_readable.xlsx (tools/fetch_ons.py)

出力は DATA_CONTRACT §2.4 の census(Ages2, 20行)/projection(20行) と
同じ shape。旧 App_Data/Population2015/Country/{ed,ep}*.json の代替。
"""
import json
import re
from pathlib import Path

import openpyxl

from . import masters

ROOT = Path(__file__).resolve().parent.parent
RAW_DIR = ROOT / "data" / "raw"

# countrycode.json のコード -> Eurostat/ONS の geo コード ("EU" だけ集計コードに読み替え)
GEO_CODES = {c: ("EU27_2020" if c == "EU" else c)
             for c in masters.COUNTRY_CODE if c != "JP"}

# Ages2 準拠の20区分 (総数を含む)。demo_pjangroup の age コードそのまま。
CENSUS_AGE_CODES = [
    "TOTAL", "Y_LT5", "Y5-9", "Y10-14", "Y15-19", "Y20-24", "Y25-29", "Y30-34",
    "Y35-39", "Y40-44", "Y45-49", "Y50-54", "Y55-59", "Y60-64", "Y65-69",
    "Y70-74", "Y75-79", "Y80-84", "Y_GE85", "UNK",
]
CENSUS_AGE_LABELS = [
    "総数", "0～4歳", "5～9歳", "10～14歳", "15～19歳", "20～24歳", "25～29歳",
    "30～34歳", "35～39歳", "40～44歳", "45～49歳", "50～54歳", "55～59歳",
    "60～64歳", "65～69歳", "70～74歳", "75～79歳", "80～84歳", "85歳以上", "年齢不詳",
]

# projection (5歳階級・90歳以上まで分離、19区分+総数=20行。旧仕様のまま)
PROJECTION_AGE_LABELS = [
    "総数", "0～4歳", "5～9歳", "10～14歳", "15～19歳", "20～24歳", "25～29歳",
    "30～34歳", "35～39歳", "40～44歳", "45～49歳", "50～54歳", "55～59歳",
    "60～64歳", "65～69歳", "70～74歳", "75～79歳", "80～84歳", "85～89歳", "90歳以上",
]

# Country(海外)の将来推計は2020年の基準値が存在しない(EUROPOP2023は2022年始まり、
# ONS 2024年基準は2024年始まり)ため、City/Pref/JP とは別に5年刻み6点で統一する。
COUNTRY_PROJECTION_YEARS = [2025, 2030, 2035, 2040, 2045, 2050]


def _flat_index(body: dict, **selectors) -> int:
    ids = body["id"]
    sizes = body["size"]
    idx = 0
    for dim_i, key in enumerate(ids):
        cat_index = body["dimension"][key]["category"]["index"]
        idx = idx * sizes[dim_i] + cat_index[selectors[key]]
    return idx


def load_census(geo: str, sex: str = "T") -> list:
    """DATA_CONTRACT §2.4 の census (Ages2, 20行×9列, 1980-2020)。

    sex: "T"(総数、既定) / "M"(男) / "F"(女)。人口ピラミッド用に "M"/"F" も使う。
    """
    by_year = json.loads((RAW_DIR / "eurostat" / "census.json").read_text(encoding="utf-8"))
    rows = [{"series": label, "population": []} for label in CENSUS_AGE_LABELS]
    for year in sorted(by_year, key=int):
        body = by_year[year]
        for i, code in enumerate(CENSUS_AGE_CODES):
            idx = _flat_index(body, freq="A", unit="NR", sex=sex, age=code, geo=geo, time=year)
            v = body["value"].get(str(idx))
            rows[i]["population"].append(int(v) if v is not None else 0)
    return rows


def _one_year_age_index(code: str) -> int | None:
    """proj_23np の1歳刻みコード ("Y_LT1","Y1".."Y99","Y_GE100") → 年齢(0-100)。"""
    if code == "Y_LT1":
        return 0
    if code == "Y_GE100":
        return 100
    m = re.fullmatch(r"Y(\d+)", code)
    return int(m.group(1)) if m else None


def load_projection_eurostat(geo: str, sex: str = "T") -> list:
    """DATA_CONTRACT §2.4 の projection (20行×6列, 2025-2050)。

    proj_23np は1歳刻みのため5歳階級に合算する (0～4歳, ..., 85～89歳, 90歳以上)。
    sex: "T"(総数、既定) / "M"(男) / "F"(女)。
    """
    by_year = json.loads((RAW_DIR / "eurostat" / "projection.json").read_text(encoding="utf-8"))
    years = sorted(by_year, key=int)
    rows = [{"series": label, "population": [0] * len(years)} for label in PROJECTION_AGE_LABELS]

    for yi, year in enumerate(years):
        body = by_year[year]
        age_index = body["dimension"]["age"]["category"]["index"]
        total_idx = _flat_index(body, freq="A", projection="BSL", sex=sex, age="TOTAL",
                                unit="PER", geo=geo, time=year)
        rows[0]["population"][yi] = int(body["value"].get(str(total_idx), 0))
        for code in age_index:
            age = _one_year_age_index(code)
            if age is None:
                continue
            bucket = 19 if age >= 90 else (age // 5 + 1)  # 90+ は行19、他は0-4→行1 ...
            idx = _flat_index(body, freq="A", projection="BSL", sex=sex, age=code,
                              unit="PER", geo=geo, time=year)
            v = body["value"].get(str(idx))
            if v is not None:
                rows[bucket]["population"][yi] += int(v)
    return rows


_ONS_AGE_RE = re.compile(r"(\d+)\s*-\s*(\d+)|(\d+)\s*and over")


_ONS_SEX = {"T": "Persons", "M": "Males", "F": "Females"}


def load_projection_uk(sex: str = "T") -> list:
    """UK は Eurostat(EUROPOP2023) 対象外のため ONS 2024-based Principal projection
    (5歳階級・年次データ) から 2025・2030・...・2050 を抜き出す。90-94/95-99/
    100-104/105以上 を合算して「90歳以上」に正規化 (DATA_CONTRACT §2.4)。
    sex: "T"(総数、既定) / "M"(男) / "F"(女)。
    """
    path = RAW_DIR / "ons" / "uk_ppp_machine_readable.xlsx"
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb["Population_in_age_groups"]
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    col_for = [header.index(str(y)) for y in COUNTRY_PROJECTION_YEARS]

    by_age = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] != _ONS_SEX[sex]:
            continue
        age_label = row[1]
        by_age[age_label] = [row[c] for c in col_for]
    wb.close()

    all_ages = list(by_age.keys())
    rows = [{"series": "総数",
             "population": [sum(by_age[a][yi] for a in all_ages) for yi in range(len(COUNTRY_PROJECTION_YEARS))]}]

    for label in PROJECTION_AGE_LABELS[1:-1]:  # "総数"と"90歳以上"を除く18区分
        lo = int(label.split("～")[0].replace("歳", ""))
        key = f"{lo} - {lo + 4}"
        rows.append({"series": label, "population": list(by_age[key])})

    old90 = ["90 - 94", "95 - 99", "100 - 104", "105 and over"]
    rows.append({"series": "90歳以上",
                "population": [sum(by_age[k][yi] for k in old90) for yi in range(len(COUNTRY_PROJECTION_YEARS))]})
    return rows
